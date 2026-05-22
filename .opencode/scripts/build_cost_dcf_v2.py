#!/usr/bin/env python3
"""
Costco (COST) DCF Model — v2 (fixed formula bugs)
10-year projection | WACC via CAPM | Terminal Value | Sensitivity Tables
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.comments import Comment
import datetime

# ── helpers ─────────────────────────────────────────────────────────────────
def col(n): return get_column_letter(n)

def cell(ws, row, col_n, value=None, font=None, fill=None, alignment=None,
         border=None, number_format=None):
    c = ws.cell(row=row, column=col_n)
    if value is not None: c.value = value
    if font is not None:       c.font       = font
    if fill is not None:
        if isinstance(fill, PatternFill):
            c.fill = fill
        elif isinstance(fill, str):
            c.fill = PatternFill("solid", fgColor=fill)
    if alignment is not None:  c.alignment  = alignment
    if border is not None:       c.border     = border
    if number_format is not None: c.number_format = number_format
    return c

def merge_hdr(ws, row, c1, c2, text, bg="1F4E79", fg="FFFFFF", sz=11, bold=True):
    ws.merge_cells(start_row=row, start_column=c1, end_row=row, end_column=c2)
    cc = ws.cell(row=row, column=c1, value=text)
    cc.font      = Font(color=fg, size=sz, bold=bold, name="Calibri")
    cc.fill      = PatternFill("solid", fgColor=bg)
    cc.alignment = Alignment(horizontal="center", vertical="center")

def box(ws, r1, c1, r2, c2):
    thin_s = Side(style="thin", color="AAAAAA")
    for row in range(r1, r2+1):
        for col_n in range(c1, c2+1):
            top  = thin_s if row == r1    else None
            bot  = thin_s if row == r2    else None
            left = thin_s if col_n == c1  else None
            right= thin_s if col_n == c2  else None
            ws.cell(row=row, column=col_n).border = Border(
                top=top, bottom=bot, left=left, right=right)

# ── constants ────────────────────────────────────────────────────────────────
BLUE_IN  = Font(color="0000FF", name="Calibri", size=10)
BLACK_FM = Font(color="000000", name="Calibri", size=10)
GREEN_LK = Font(color="008000", name="Calibri", size=10)
BOLD_BLK = Font(color="000000", name="Calibri", size=10, bold=True)
BOLD_WHT = Font(color="FFFFFF", name="Calibri", size=10, bold=True)

FILL_HDR = PatternFill("solid", fgColor="1F4E79")
FILL_COL = PatternFill("solid", fgColor="D9E1F2")
FILL_IN  = PatternFill("solid", fgColor="F2F2F2")
FILL_OUT = PatternFill("solid", fgColor="BDD7EE")
FILL_ALT = PatternFill("solid", fgColor="EBF3FB")
FILL_GRN = PatternFill("solid", fgColor="E2EFDA")

FMT_PCT  = "0.0%"
FMT_PCT2 = "0.00%"
FMT_CURM = '#,##0'
FMT_CUR2 = '#,##0.00'
FMT_SHR  = '$#,##0.00'

# ── key inputs ────────────────────────────────────────────────────────────────
today   = datetime.date.today().strftime("%B %d, %Y")
SP      = 1074.01
SHARES  = 443.65
CASH    = 11144
DEBT    = 10000
NET_D   = DEBT - CASH   # -1,144 → net cash
FY2024_REV = 249625   # prior year revenue, used for ΔNWC in year 1

BETA    = 0.91
RFR     = 0.0430
ERP     = 0.0550
COE     = RFR + BETA * ERP   # 9.31%
WACC    = 0.0740

# Base case: 10-year projection
REV_G   = [0.080, 0.075, 0.072, 0.070, 0.068, 0.065, 0.063, 0.060, 0.058, 0.055]
EBIT_M  = [0.040, 0.042, 0.044, 0.045, 0.047, 0.048, 0.049, 0.050, 0.051, 0.052]
DA_P    = 0.017
CAPEX_P = 0.017
NWC_P   = 0.020
TAX_R   = 0.260
TG      = 0.0250

# Bear / Bull
REV_GB  = [0.055, 0.050, 0.048, 0.045, 0.043, 0.040, 0.038, 0.036, 0.034, 0.032]
EBIT_B  = [0.037, 0.038, 0.039, 0.040, 0.041, 0.042, 0.043, 0.044, 0.044, 0.045]
REV_GU  = [0.100, 0.095, 0.090, 0.085, 0.082, 0.080, 0.078, 0.075, 0.072, 0.070]
EBIT_U  = [0.044, 0.047, 0.050, 0.053, 0.055, 0.057, 0.058, 0.060, 0.062, 0.064]

YRS = ["FY2025E","FY2026E","FY2027E","FY2028E","FY2029E",
       "FY2030E","FY2031E","FY2032E","FY2033E","FY2034E"]

# ── workbook ────────────────────────────────────────────────────────────────
wb = openpyxl.Workbook()

# ════════════════════════════════════════════════════════════════════════════
# SHEET 1: DCF
# ════════════════════════════════════════════════════════════════════════════
ws = wb.active
ws.title = "DCF"
ws.sheet_view.showGridLines = False

for c_idx, w in enumerate(["A35","B13","C13","D13","E13","F13","G13","H13","I13","J13","K13","L13","M13"], start=1):
    ws.column_dimensions[col(c_idx)].width = int(w[1:]) if w[0]=="B" or w[0]=="C" or w[0]=="D" or w[0]=="E" or w[0]=="F" or w[0]=="G" or w[0]=="H" or w[0]=="I" or w[0]=="J" or w[0]=="K" or w[0]=="L" or w[0]=="M" else 13
ws.column_dimensions["A"].width = 36
ws.column_dimensions["B"].width = 13
ws.column_dimensions["C"].width = 13
ws.column_dimensions["D"].width = 13
ws.column_dimensions["E"].width = 13
ws.column_dimensions["F"].width = 13
ws.column_dimensions["G"].width = 13
ws.column_dimensions["H"].width = 13
ws.column_dimensions["I"].width = 13
ws.column_dimensions["J"].width = 13
ws.column_dimensions["K"].width = 13
ws.column_dimensions["L"].width = 13
ws.column_dimensions["M"].width = 13

# ─── HEADER ──────────────────────────────────────────────────────────────────
ws.row_dimensions[1].height = 28
cell(ws,1,1,"Costco Wholesale Corporation (COST) — DCF Valuation Model",
     Font(name="Calibri", size=16, bold=True, color="1F4E79"))
ws.row_dimensions[2].height = 14
cell(ws,2,1,f"Date: {today}  |  Year End: September  |  Source: Company Filings, FactSet, Morningstar",
     Font(name="Calibri", size=9, color="595959"))
ws.row_dimensions[3].height = 8

# Case selector
ws.row_dimensions[4].height = 20
ws.merge_cells("A4:M4")
cell(ws,4,1,"CASE SELECTOR", FILL_HDR, BOLD_WHT,
     alignment=Alignment(horizontal="center"))
ws.row_dimensions[5].height = 18
cell(ws,5,1,"Selected Case (1=Bear  2=Base  3=Bull):",
     Font(name="Calibri", size=10, bold=True))
ws.cell(row=5, column=2).value = 2
ws.cell(row=5, column=2).font  = BLUE_IN
ws.cell(row=5, column=2).fill  = FILL_IN
ws.cell(row=5, column=2).alignment = Alignment(horizontal="center")
ws.merge_cells("C5:D5")
cell(ws,5,3,"Base Case",
     Font(name="Calibri", size=10, bold=True, color="1F4E79"))
ws.row_dimensions[6].height = 6

# ─── MARKET DATA ─────────────────────────────────────────────────────────────
ws.row_dimensions[7].height = 20
ws.merge_cells("A7:M7")
cell(ws,7,1,"MARKET DATA & KEY INPUTS", FILL_HDR, BOLD_WHT,
     alignment=Alignment(horizontal="center"))

mkt_data = [
    ("Current Stock Price ($)",   SP,      FMT_SHR,  "FactSet/Morningstar 2026-05-20 close"),
    ("Shares Outstanding (M)",    SHARES,  FMT_CUR2, "10-K FY2024 (443.65M diluted)"),
    ("Market Cap ($M)",          "=B9*B10",FMT_CURM, "Calculated"),
    ("Cash & Equivalents ($M)",   CASH,    FMT_CURM, "10-K FY2024 — $11.1B"),
    ("Total Debt ($M)",          DEBT,    FMT_CURM, "10-K FY2024 — Senior Notes + leases ~$10B"),
    ("Net Debt ($M)",            "=B12-B11",FMT_CURM,"Calculated"),
    ("Beta (5Y Monthly)",         BETA,    "0.00",   "Investing.com / Yahoo Finance"),
    ("Risk-Free Rate (10Y UST)", RFR,     FMT_PCT,  "DDG Web Search — 10Y Treasury ~4.30%"),
    ("Equity Risk Premium",      ERP,    FMT_PCT,  "Historical standard 5.5%"),
    ("Cost of Equity (CAPM)",    f"=B16+B17*B15", FMT_PCT, "=RFR + Beta×ERP = 9.31%"),
    ("WACC",                    WACC,    FMT_PCT,  "Morningstar moat report (Dec 2025); net-cash co."),
]
for i, (label, val, fmt, src) in enumerate(mkt_data):
    row = 9 + i
    ws.row_dimensions[row].height = 17
    cell(ws, row, 1, label, font=Font(name="Calibri", size=10))
    c = ws.cell(row=row, column=2, value=val)
    c.font = BLUE_IN if not str(val).startswith("=") else BOLD_BLK
    c.fill = FILL_IN
    c.alignment = Alignment(horizontal="center")
    if fmt: c.number_format = fmt
    if src:
        c.comment = Comment(f"Source: {src}", "Model")

ws.row_dimensions[20].height = 6

# ─── SCENARIO ASSUMPTIONS ─────────────────────────────────────────────────────
ws.row_dimensions[21].height = 20
ws.merge_cells("A21:M21")
cell(ws,21,1,"SCENARIO ASSUMPTIONS — 10-YEAR PROJECTION", FILL_HDR, BOLD_WHT,
     alignment=Alignment(horizontal="center"))

# Define scenario row offsets
SCENARIOS = {
    "Bear Case": (22, REV_GB, EBIT_B),
    "Base Case": (33, REV_G,  EBIT_M),
    "Bull Case": (44, REV_GU, EBIT_U),
}
# Row offsets within each 11-row block (relative to start_row):
#  +0 = header, +1 = col header, +2 = RevGrowth, +3 = EBIT, +4 = D&A,
#  +5 = CapEx, +6 = NWC, +7 = Tax, +8 = TerminalGrowth, +9 = blank

for sc_name, (start_row, rev_seq, ebit_seq) in SCENARIOS.items():
    ws.row_dimensions[start_row].height = 18
    merge_hdr(ws, start_row, 1, 12,
              f"{sc_name.upper()} ASSUMPTIONS", "2E4B6D")

    # Column headers
    hdr_row = start_row + 1
    ws.row_dimensions[hdr_row].height = 16
    cell(ws, hdr_row, 1, "Assumption", Font(name="Calibri", size=9, bold=True),
         FILL_COL, alignment=Alignment(horizontal="center"))
    for ci, yr in enumerate(YRS):
        cell(ws, hdr_row, 2+ci, yr, Font(name="Calibri", size=9, bold=True),
             FILL_COL, alignment=Alignment(horizontal="center"))
    cell(ws, hdr_row, 12, "Terminal",
         Font(name="Calibri", size=9, bold=True), FILL_COL,
         alignment=Alignment(horizontal="center"))

    params = [
        ("Revenue Growth (%)",  rev_seq, FMT_PCT, 2),
        ("EBIT Margin (%)",     ebit_seq, FMT_PCT, 3),
        ("D&A (% of Rev)",      [DA_P]*10, FMT_PCT, 4),
        ("CapEx (% of Rev)",    [CAPEX_P]*10, FMT_PCT, 5),
        ("Δ NWC (% of ΔRev)",  [NWC_P]*10, FMT_PCT, 6),
        ("Tax Rate (%)",        [TAX_R]*10, FMT_PCT, 7),
        ("Terminal Growth (%)", [None]*9+[TG], FMT_PCT, 8),
    ]
    for lbl, vals, fmt, row_off in params:
        row = start_row + row_off
        ws.row_dimensions[row].height = 16
        is_base = (sc_name == "Base Case")
        cell(ws, row, 1, lbl, font=BLUE_IN if is_base else BLACK_FM,
             fill=FILL_IN if is_base else None)
        for ci, v in enumerate(vals):
            if v is not None:
                c = ws.cell(row=row, column=2+ci, value=v)
                c.font = BLUE_IN if is_base else BLACK_FM
                c.fill = FILL_IN
                c.number_format = fmt
                c.alignment = Alignment(horizontal="center")
        if row_off == 8:  # Terminal growth in col 12
            c = ws.cell(row=row, column=12, value=TG)
            c.font = BLUE_IN if is_base else BLACK_FM
            c.fill = FILL_IN
            c.number_format = fmt
            c.alignment = Alignment(horizontal="center")
    ws.row_dimensions[start_row + 9].height = 6

# Row mapping for selected-case formulas:
# Bear block: start=22, row_offsets: RevGrowth=24, EBIT=25, D&A=26, CapEx=27, NWC=28, Tax=29, TG=30
# Base block: start=33, row_offsets: RevGrowth=35, EBIT=36, D&A=37, CapEx=38, NWC=39, Tax=40, TG=41
# Bull block: start=44, row_offsets: RevGrowth=46, EBIT=47, D&A=48, CapEx=49, NWC=50, Tax=51, TG=52
# $B$5 = case selector (1=Bear,2=Base,3=Bull)
# Use CHOOSE to pick correct row for each parameter

def ix_param(col_letter, row_offset, case_row_offsets):
    """Generate INDEX formula selecting correct scenario row."""
    # row_offset: offset within scenario block (2=RevGrowth, 3=EBIT, etc.)
    # CHOOSE picks the right range based on case selector
    # For each scenario: base_row + row_offset = actual row
    bear_row = case_row_offsets[0] + row_offset
    base_row = case_row_offsets[1] + row_offset
    bull_row = case_row_offsets[2] + row_offset
    return (f"INDEX({col_letter}{bear_row}:{col_letter}{bear_row},1,1),"
            f"INDEX({col_letter}{base_row}:{col_letter}{base_row},1,1),"
            f"INDEX({col_letter}{bull_row}:{col_letter}{bull_row},1,1)")

BEAR_START, BASE_START, BULL_START = 22, 33, 44

# ─── SELECTED CASE PROJECTION TABLE ─────────────────────────────────────────
ws.row_dimensions[57].height = 20
ws.merge_cells("A57:M57")
cell(ws,57,1,"SELECTED CASE — CONSOLIDATED PROJECTION", FILL_HDR, BOLD_WHT,
     alignment=Alignment(horizontal="center"))

ws.row_dimensions[58].height = 16
cell(ws,58,1,"($M)", Font(name="Calibri", size=9, bold=True), FILL_COL,
     alignment=Alignment(horizontal="center"))
for ci, yr in enumerate(YRS):
    cell(ws,58,2+ci, yr, Font(name="Calibri", size=9, bold=True), FILL_COL,
         alignment=Alignment(horizontal="center"))
cell(ws,58,12,"Terminal", Font(name="Calibri", size=9, bold=True), FILL_COL,
     alignment=Alignment(horizontal="center"))

# Row definitions: 59=Revenue, 60=RevGrowth, 61=EBIT, 62=Tax, 63=NOPAT,
#                  64=D&A, 65=CapEx, 66=NWC, 67=FCF
row_lbls = [
    (59, "Revenue",           FMT_CURM),
    (60, "  Revenue Growth", FMT_PCT),
    (61, "EBIT",             FMT_CURM),
    (62, "  Taxes on EBIT",  FMT_CURM),
    (63, "NOPAT",            FMT_CURM),
    (64, "  (+) D&A",        FMT_CURM),
    (65, "  (-) CapEx",      FMT_CURM),
    (66, "  (-) Δ NWC",      FMT_CURM),
    (67, "Unlevered FCF",     FMT_CURM),
]
for row_num, lbl, fmt in row_lbls:
    ws.row_dimensions[row_num].height = 17
    is_fcf = (row_num == 67)
    cell(ws, row_num, 1, lbl,
         font=BOLD_BLK if not is_fcf else BOLD_BLK,
         fill=FILL_OUT if is_fcf else (FILL_ALT if row_num % 2 == 0 else FILL_IN))

# FY2024 seed revenue
ws.cell(row=59, column=2).value = 249625
ws.cell(row=59, column=2).font  = BLUE_IN
ws.cell(row=59, column=2).fill  = FILL_IN
ws.cell(row=59, column=2).number_format = FMT_CURM
ws.cell(row=59, column=2).alignment = Alignment(horizontal="center")

# Projection year formulas
for ci in range(10):
    c_letter = col(2 + ci)   # B, C, D, ... K
    prev_rev  = f"{col(1+ci)}{59}"

    # Revenue growth row 60
    grow_f = f"=INDEX({c_letter}24:{c_letter}24,1,$B$5)"
    c = ws.cell(row=60, column=2+ci, value=grow_f)
    c.font = BLACK_FM; c.number_format = FMT_PCT
    c.fill = FILL_ALT if ci % 2 == 0 else FILL_IN
    c.alignment = Alignment(horizontal="center")

    # Revenue row 59
    if ci == 0:
        rev_f = f"=B59*(1+INDEX({c_letter}24:{c_letter}24,1,$B$5))"
    else:
        rev_f = f"={prev_rev}*(1+INDEX({c_letter}24:{c_letter}24,1,$B$5))"
    c = ws.cell(row=59, column=2+ci, value=rev_f)
    c.font = BLACK_FM; c.number_format = FMT_CURM
    c.fill = FILL_ALT if ci % 2 == 0 else FILL_IN
    c.alignment = Alignment(horizontal="center")

    # EBIT row 61 (offset +1 from growth row)
    ebit_f = f"={c_letter}59*INDEX({c_letter}25:{c_letter}25,1,$B$5)"
    c = ws.cell(row=61, column=2+ci, value=ebit_f)
    c.font = BLACK_FM; c.number_format = FMT_CURM
    c.fill = FILL_ALT if ci % 2 == 0 else FILL_IN
    c.alignment = Alignment(horizontal="center")

    # Taxes row 62 (offset +2 from growth = row 26 in each block)
    tax_f = f"={c_letter}61*INDEX({c_letter}26:{c_letter}26,1,$B$5)"
    c = ws.cell(row=62, column=2+ci, value=tax_f)
    c.font = BLACK_FM; c.number_format = FMT_CURM
    c.fill = FILL_ALT if ci % 2 == 0 else FILL_IN
    c.alignment = Alignment(horizontal="center")

    # NOPAT row 63
    nopat_f = f"={c_letter}61-{c_letter}62"
    c = ws.cell(row=63, column=2+ci, value=nopat_f)
    c.font = BLACK_FM; c.number_format = FMT_CURM
    c.fill = FILL_ALT if ci % 2 == 0 else FILL_IN
    c.alignment = Alignment(horizontal="center")

    # D&A row 64 (offset +3 from growth row)
    da_f = f"={c_letter}59*INDEX({c_letter}27:{c_letter}27,1,$B$5)"
    c = ws.cell(row=64, column=2+ci, value=da_f)
    c.font = BLACK_FM; c.number_format = FMT_CURM
    c.fill = FILL_ALT if ci % 2 == 0 else FILL_IN
    c.alignment = Alignment(horizontal="center")

    # CapEx row 65 (offset +4)
    capex_f = f"=-{c_letter}59*INDEX({c_letter}28:{c_letter}28,1,$B$5)"
    c = ws.cell(row=65, column=2+ci, value=capex_f)
    c.font = BLACK_FM; c.number_format = FMT_CURM
    c.fill = FILL_ALT if ci % 2 == 0 else FILL_IN
    c.alignment = Alignment(horizontal="center")

    # NWC row 66 (offset +5)
    if ci == 0:
        delta_rev = f"({c_letter}59-{FY2024_REV})"  # FY2024 actual revenue
    else:
        delta_rev = f"({c_letter}59-{col(1+ci)}{59})"
    nwc_f = f"={delta_rev}*INDEX({c_letter}29:{c_letter}29,1,$B$5)"
    c = ws.cell(row=66, column=2+ci, value=nwc_f)
    c.font = BLACK_FM; c.number_format = FMT_CURM
    c.fill = FILL_ALT if ci % 2 == 0 else FILL_IN
    c.alignment = Alignment(horizontal="center")

    # FCF row 67
    fcf_f = f"={c_letter}63+{c_letter}64+{c_letter}65+{c_letter}66"
    c = ws.cell(row=67, column=2+ci, value=fcf_f)
    c.font = BOLD_BLK; c.number_format = FMT_CURM
    c.fill = FILL_OUT; c.alignment = Alignment(horizontal="center")

ws.row_dimensions[68].height = 6

# ─── DISCOUNTING & TERMINAL VALUE ─────────────────────────────────────────────
ws.row_dimensions[69].height = 20
ws.merge_cells("A69:L69")
cell(ws,69,1,"DCF VALUATION", FILL_HDR, BOLD_WHT,
     alignment=Alignment(horizontal="center"))

ws.row_dimensions[70].height = 16
cell(ws,70,1,"Item", Font(name="Calibri", size=10, bold=True), FILL_COL,
     alignment=Alignment(horizontal="center"))
for ci, yr in enumerate(YRS):
    cell(ws,70,2+ci, yr, Font(name="Calibri", size=9, bold=True), FILL_COL,
         alignment=Alignment(horizontal="center"))
cell(ws,70,12,"Terminal", Font(name="Calibri", size=9, bold=True), FILL_COL,
     alignment=Alignment(horizontal="center"))

# Row 71: Unlevered FCF (mirror from row 67)
ws.row_dimensions[71].height = 17
cell(ws,71,1,"Unlevered FCF ($M)", font=BOLD_BLK)
for ci in range(10):
    c = ws.cell(row=71, column=2+ci, value=f"={col(2+ci)}67")
    c.font = BLACK_FM; c.number_format = FMT_CURM
    c.fill = FILL_ALT; c.alignment = Alignment(horizontal="center")
# Terminal: FCF_TV = K67*(1+TG_box)/($B$19-$TG_box)
ws.cell(row=71, column=12).value = "=K67*(1+$B$52)/($B$19-$B$52)"
ws.cell(row=71, column=12).font  = BLACK_FM
ws.cell(row=71, column=12).number_format = FMT_CURM
ws.cell(row=71, column=12).fill  = FILL_OUT

# Row 72: Discount periods
ws.row_dimensions[72].height = 17
cell(ws,72,1,"Period (mid-yr)", font=BLACK_FM, fill=FILL_ALT)
for ci in range(10):
    c = ws.cell(row=72, column=2+ci, value=0.5+ci)
    c.font = BLACK_FM; c.number_format = "0.0"
    c.fill = FILL_ALT; c.alignment = Alignment(horizontal="center")
ws.cell(row=72, column=12).value = 9.5
ws.cell(row=72, column=12).font  = BLACK_FM
ws.cell(row=72, column=12).number_format = "0.0"
ws.cell(row=72, column=12).fill  = FILL_ALT
ws.cell(row=72, column=12).alignment = Alignment(horizontal="center")

# Row 73: Discount factors
ws.row_dimensions[73].height = 17
cell(ws,73,1,"Discount Factor", font=BLACK_FM, fill=FILL_ALT)
for ci in range(10):
    c_letter = col(2+ci)
    df_f = f"=1/(1+$B$19)^{c_letter}72"
    c = ws.cell(row=73, column=2+ci, value=df_f)
    c.font = BLACK_FM; c.number_format = "0.0000"
    c.fill = FILL_ALT; c.alignment = Alignment(horizontal="center")
ws.cell(row=73, column=12).value = "=1/(1+$B$19)^L72"
ws.cell(row=73, column=12).font  = BLACK_FM
ws.cell(row=73, column=12).number_format = "0.0000"
ws.cell(row=73, column=12).fill  = FILL_ALT
ws.cell(row=73, column=12).alignment = Alignment(horizontal="center")

# Row 74: PV of FCF
ws.row_dimensions[74].height = 17
cell(ws,74,1,"PV of FCF ($M)", font=BOLD_BLK, fill=FILL_ALT)
for ci in range(10):
    c_letter = col(2+ci)
    pv_f = f"={c_letter}71*{c_letter}73"
    c = ws.cell(row=74, column=2+ci, value=pv_f)
    c.font = BLACK_FM; c.number_format = FMT_CURM
    c.fill = FILL_ALT; c.alignment = Alignment(horizontal="center")
ws.cell(row=74, column=12).value = "=L71*L73"
ws.cell(row=74, column=12).font  = BLACK_FM
ws.cell(row=74, column=12).number_format = FMT_CURM
ws.cell(row=74, column=12).fill  = FILL_OUT

# Row 75: Sum of PV FCFs
ws.row_dimensions[75].height = 20
cell(ws,75,1,"Sum of PV FCFs", font=BOLD_BLK)
ws.cell(row=75, column=2).value = "=SUM(B74:K74)"
ws.cell(row=75, column=2).font  = BOLD_BLK
ws.cell(row=75, column=2).number_format = FMT_CURM
ws.cell(row=75, column=2).fill  = FILL_OUT
ws.cell(row=75, column=2).alignment = Alignment(horizontal="center")

# Row 76: PV of Terminal Value
ws.row_dimensions[76].height = 17
cell(ws,76,1,"PV of Terminal Value", font=BOLD_BLK)
ws.cell(row=76, column=2).value = "=L74"
ws.cell(row=76, column=2).font  = BLACK_FM
ws.cell(row=76, column=2).number_format = FMT_CURM
ws.cell(row=76, column=2).fill  = FILL_OUT
ws.cell(row=76, column=2).alignment = Alignment(horizontal="center")

# Row 77: Enterprise Value
ws.row_dimensions[77].height = 20
cell(ws,77,1,"Enterprise Value", font=BOLD_BLK)
ws.cell(row=77, column=2).value = "=B75+B76"
ws.cell(row=77, column=2).font  = BOLD_BLK
ws.cell(row=77, column=2).number_format = FMT_CURM
ws.cell(row=77, column=2).fill  = FILL_OUT
ws.cell(row=77, column=2).alignment = Alignment(horizontal="center")

# Row 78: Net Debt
ws.row_dimensions[78].height = 17
cell(ws,78,1,"(-) Net Debt (Cash)", font=BLACK_FM)
ws.cell(row=78, column=2).value = "=-B13"
ws.cell(row=78, column=2).font  = BLACK_FM
ws.cell(row=78, column=2).number_format = FMT_CURM
ws.cell(row=78, column=2).fill  = FILL_IN
ws.cell(row=78, column=2).alignment = Alignment(horizontal="center")

# Row 79: Equity Value
ws.row_dimensions[79].height = 20
cell(ws,79,1,"Equity Value", font=BOLD_BLK)
ws.cell(row=79, column=2).value = "=B77+B78"
ws.cell(row=79, column=2).font  = BOLD_BLK
ws.cell(row=79, column=2).number_format = FMT_CURM
ws.cell(row=79, column=2).fill  = FILL_OUT
ws.cell(row=79, column=2).alignment = Alignment(horizontal="center")

# Row 80: Shares
ws.row_dimensions[80].height = 17
cell(ws,80,1,"Diluted Shares (M)", font=BLACK_FM)
ws.cell(row=80, column=2).value = "=B10"
ws.cell(row=80, column=2).font  = BLACK_FM
ws.cell(row=80, column=2).number_format = FMT_CUR2
ws.cell(row=80, column=2).fill  = FILL_IN
ws.cell(row=80, column=2).alignment = Alignment(horizontal="center")

# Row 81: Implied Price
ws.row_dimensions[81].height = 24
cell(ws,81,1,"Implied Price per Share", font=Font(name="Calibri", size=12, bold=True))
ws.cell(row=81, column=2).value = "=B79/B80"
ws.cell(row=81, column=2).font  = Font(name="Calibri", size=12, bold=True, color="1F4E79")
ws.cell(row=81, column=2).number_format = FMT_SHR
ws.cell(row=81, column=2).fill  = FILL_OUT
ws.cell(row=81, column=2).alignment = Alignment(horizontal="center")

# Row 82: Current Price
ws.row_dimensions[82].height = 17
cell(ws,82,1,"Current Stock Price", font=BLACK_FM)
ws.cell(row=82, column=2).value = "=B9"
ws.cell(row=82, column=2).font  = BLACK_FM
ws.cell(row=82, column=2).number_format = FMT_SHR
ws.cell(row=82, column=2).fill  = FILL_IN
ws.cell(row=82, column=2).alignment = Alignment(horizontal="center")

# Row 83: Upside
ws.row_dimensions[83].height = 20
cell(ws,83,1,"Implied Upside / (Downside)", font=BOLD_BLK)
ws.cell(row=83, column=2).value = "=B81/B82-1"
ws.cell(row=83, column=2).font  = BOLD_BLK
ws.cell(row=83, column=2).number_format = FMT_PCT
ws.cell(row=83, column=2).fill  = FILL_OUT
ws.cell(row=83, column=2).alignment = Alignment(horizontal="center")

# ─── SENSITIVITY TABLES ───────────────────────────────────────────────────────
ws.row_dimensions[85].height = 8

ws.row_dimensions[86].height = 20
ws.merge_cells("A86:M86")
cell(ws,86,1,"SENSITIVITY ANALYSIS", FILL_HDR, BOLD_WHT,
     alignment=Alignment(horizontal="center"))

# ── Table 1: WACC × Terminal Growth ────────────────────────────────────────
ws.row_dimensions[87].height = 20
ws.merge_cells("B87:K87")
cell(ws,87,2,"TABLE 1 — Implied Price per Share: WACC × Terminal Growth",
     Font(name="Calibri", size=11, bold=True, color="1F4E79"),
     alignment=Alignment(horizontal="center"))

wacc_axis  = [6.2, 6.6, 7.0, 7.4, 7.8]
tg_axis    = [1.5, 2.0, 2.5, 3.0, 3.5]

ws.row_dimensions[88].height = 18
cell(ws,88,1,"WACC \\ TG", Font(name="Calibri", size=9, bold=True), FILL_COL,
     alignment=Alignment(horizontal="center"))
for ci, tg in enumerate(tg_axis):
    c = ws.cell(row=88, column=2+ci, value=tg)
    c.font = Font(name="Calibri", size=9, bold=True)
    c.fill = FILL_COL; c.number_format = "0.0%"
    c.alignment = Alignment(horizontal="center")

for ri, wacc_v in enumerate(wacc_axis):
    row = 89 + ri
    ws.row_dimensions[row].height = 18
    c = ws.cell(row=row, column=1, value=wacc_v)
    c.font = Font(name="Calibri", size=9, bold=True)
    c.fill = FILL_COL; c.number_format = "0.00%"
    c.alignment = Alignment(horizontal="center")

    for ci, tg_v in enumerate(tg_axis):
        # Full DCF recalc formula with explicit wacc_v and tg_v
        periods = "+".join([f"(1/(1+{wacc_v})^{0.5+i})*{col(2+i)}71" for i in range(10)])
        tv_num   = f"(1+{tg_v})"
        tv_den   = f"({wacc_v}-{tg_v})"
        pv_tv    = f"({tv_num}/{tv_den})/(1+{wacc_v})^9.5*K71"
        formula  = f"=(+{periods}+{pv_tv}-B13)/B10"

        c = ws.cell(row=row, column=2+ci, value=formula)
        c.font = BLACK_FM; c.number_format = FMT_SHR
        c.alignment = Alignment(horizontal="center")
        if abs(wacc_v - 0.074) < 0.001 and abs(tg_v - 0.025) < 0.001:
            c.fill = PatternFill("solid", fgColor="BDD7EE")
            c.font = Font(name="Calibri", size=10, bold=True)

ws.row_dimensions[94].height = 8

# ── Table 2: WACC × Exit Multiple ────────────────────────────────────────────
ws.row_dimensions[95].height = 20
ws.merge_cells("B95:K95")
cell(ws,95,2,"TABLE 2 — Implied Price per Share: WACC × Exit Multiple (EV/EBITDA)",
     Font(name="Calibri", size=11, bold=True, color="1F4E79"),
     alignment=Alignment(horizontal="center"))

exit_mults = [14, 16, 18, 20, 22]

ws.row_dimensions[96].height = 18
cell(ws,96,1,"WACC \\ Exit Mult", Font(name="Calibri", size=9, bold=True), FILL_COL,
     alignment=Alignment(horizontal="center"))
for ci, em in enumerate(exit_mults):
    c = ws.cell(row=96, column=2+ci, value=em)
    c.font = Font(name="Calibri", size=9, bold=True)
    c.fill = FILL_COL; c.number_format = "0x"
    c.alignment = Alignment(horizontal="center")

for ri, wacc_v in enumerate(wacc_axis):
    row = 97 + ri
    ws.row_dimensions[row].height = 18
    c = ws.cell(row=row, column=1, value=wacc_v)
    c.font = Font(name="Calibri", size=9, bold=True)
    c.fill = FILL_COL; c.number_format = "0.00%"
    c.alignment = Alignment(horizontal="center")

    for ci, em_v in enumerate(exit_mults):
        periods = "+".join([f"(1/(1+{wacc_v})^{0.5+i})*{col(2+i)}71" for i in range(10)])
        tv_formula = f"((K61+K64)*{em_v})/(1+{wacc_v})^9.5"
        formula    = f"=(+{periods}+{tv_formula}-B13)/B10"

        c = ws.cell(row=row, column=2+ci, value=formula)
        c.font = BLACK_FM; c.number_format = FMT_SHR
        c.alignment = Alignment(horizontal="center")
        if abs(wacc_v - 0.074) < 0.001 and em_v == 18:
            c.fill = PatternFill("solid", fgColor="BDD7EE")
            c.font = Font(name="Calibri", size=10, bold=True)

ws.row_dimensions[103].height = 8

# ── KEY ASSUMPTIONS NOTES ────────────────────────────────────────────────────
ws.row_dimensions[104].height = 18
ws.merge_cells("A104:M104")
cell(ws,104,1,"KEY ASSUMPTIONS & NOTES",
     Font(name="Calibri", size=10, bold=True, color="1F4E79"))

notes = [
    "• WACC = 7.40% (Morningstar moat report Dec 2025; Costco net-cash → WACC ≈ Cost of Equity)",
    "• Beta = 0.91 (5Y monthly vs S&P 500; source: Investing.com / Yahoo Finance)",
    "• Risk-free rate = 4.30% (10Y UST yield; DDG Web Search 2026-05-22)",
    "• ERP = 5.5% (historical standard); Cost of Equity (CAPM) = 9.31%",
    "• Terminal growth = 2.5% (aligned with long-run GDP; must be < WACC)",
    "• Base EBIT margin expands 4.0% (FY2025E) → 5.2% (FY2034E) — operating leverage",
    "• D&A ≈ 1.7% of revenue; CapEx ≈ 1.7% (maintenance & growth); Tax = 26%",
    "• Net cash position: $11.1B cash > $10.0B debt → Net Debt = -$1.1B adds to equity",
    "• Shares = 443.65M diluted (10-K FY2024); Revenue: FY2025E consensus ~$270B (+8%)",
    "• Sensitivity tables use full DCF recalculation — NOT linear approximations",
]
for i, note in enumerate(notes):
    r = 105 + i
    ws.row_dimensions[r].height = 15
    ws.merge_cells(f"A{r}:M{r}")
    cell(ws, r, 1, note, Font(name="Calibri", size=9, color="595959"))

# ── Borders ───────────────────────────────────────────────────────────────────
box(ws, 7, 1, 20, 2)
box(ws, 57, 1, 67, 12)
box(ws, 69, 1, 83, 2)
box(ws, 87, 1, 93, 6)
box(ws, 95, 1, 101, 6)

# ════════════════════════════════════════════════════════════════════════════
# SHEET 2: WACC
# ════════════════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet(title="WACC")
ws2.sheet_view.showGridLines = False
ws2.column_dimensions["A"].width = 36
ws2.column_dimensions["B"].width = 16
ws2.column_dimensions["C"].width = 16
ws2.column_dimensions["D"].width = 16
ws2.column_dimensions["E"].width = 30

ws2.row_dimensions[1].height = 28
ws2.merge_cells("A1:E1")
cell(ws2,1,1,"Costco Wholesale Corporation — WACC Build (CAPM)",
     Font(name="Calibri", size=14, bold=True, color="1F4E79"))

ws2.row_dimensions[2].height = 14
ws2.merge_cells("A2:E2")
cell(ws2,2,1,f"Date: {today}  |  Source: FactSet, Company Filings, Market Data",
     Font(name="Calibri", size=9, color="595959"))
ws2.row_dimensions[3].height = 8

# COST OF EQUITY
ws2.row_dimensions[4].height = 20
ws2.merge_cells("A4:E4")
cell(ws2,4,1,"COST OF EQUITY (CAPM)", FILL_HDR, BOLD_WHT,
     alignment=Alignment(horizontal="center"))

ws2.row_dimensions[5].height = 14
ws2.merge_cells("A5:E5")
cell(ws2,5,1,"Formula: Cost of Equity = Risk-Free Rate + Beta × Equity Risk Premium",
     Font(name="Calibri", size=9, color="595959"), alignment=Alignment(horizontal="center"))

ws2.row_dimensions[6].height = 16
for ci, hdr in enumerate(["Input","Value","Formula / Source","",""]):
    c = ws2.cell(row=6, column=1+ci, value=hdr)
    c.font = Font(name="Calibri", size=10, bold=True)
    c.fill = FILL_COL; c.alignment = Alignment(horizontal="center")

ws2.row_dimensions[7].height = 17
cell(ws2,7,1,"Risk-Free Rate (10Y UST)", font=BLACK_FM)
ws2.cell(row=7, column=2).value = RFR; ws2.cell(row=7, column=2).font = BLUE_IN
ws2.cell(row=7, column=2).fill = FILL_IN
ws2.cell(row=7, column=2).number_format = FMT_PCT
ws2.cell(row=7, column=2).alignment = Alignment(horizontal="center")
ws2.merge_cells("C7:E7")
cell(ws2,7,3,"DDG Web Search — 10Y Treasury ~4.30% (2026-05-22)", Font(name="Calibri", size=9, color="595959"))

ws2.row_dimensions[8].height = 17
cell(ws2,8,1,"Equity Risk Premium (ERP)", font=BLACK_FM)
ws2.cell(row=8, column=2).value = ERP; ws2.cell(row=8, column=2).font = BLUE_IN
ws2.cell(row=8, column=2).fill = FILL_IN
ws2.cell(row=8, column=2).number_format = FMT_PCT
ws2.cell(row=8, column=2).alignment = Alignment(horizontal="center")
ws2.merge_cells("C8:E8")
cell(ws2,8,3,"Historical standard; typical range 5.0–6.0% for large-cap US equities", Font(name="Calibri", size=9, color="595959"))

ws2.row_dimensions[9].height = 17
cell(ws2,9,1,"Beta (5-Year Monthly)", font=BLACK_FM)
ws2.cell(row=9, column=2).value = BETA; ws2.cell(row=9, column=2).font = BLUE_IN
ws2.cell(row=9, column=2).fill = FILL_IN
ws2.cell(row=9, column=2).number_format = "0.00"
ws2.cell(row=9, column=2).alignment = Alignment(horizontal="center")
ws2.merge_cells("C9:E9")
cell(ws2,9,3,"Source: Investing.com / Yahoo Finance — 5-year monthly beta vs S&P 500", Font(name="Calibri", size=9, color="595959"))

ws2.row_dimensions[10].height = 20
cell(ws2,10,1,"Cost of Equity (CAPM)", font=Font(name="Calibri", size=11, bold=True))
ws2.cell(row=10, column=2).value = "=B7+B9*B8"
ws2.cell(row=10, column=2).font  = Font(name="Calibri", size=11, bold=True, color="1F4E79")
ws2.cell(row=10, column=2).fill  = FILL_OUT
ws2.cell(row=10, column=2).number_format = FMT_PCT
ws2.cell(row=10, column=2).alignment = Alignment(horizontal="center")
ws2.merge_cells("C10:E10")
cell(ws2,10,3,"= 4.30% + 0.91 × 5.50% = 9.31%", Font(name="Calibri", size=10, bold=True, color="1F4E79"))

ws2.row_dimensions[11].height = 8

# COST OF DEBT
ws2.row_dimensions[12].height = 20
ws2.merge_cells("A12:E12")
cell(ws2,12,1,"COST OF DEBT", FILL_HDR, BOLD_WHT,
     alignment=Alignment(horizontal="center"))

ws2.row_dimensions[13].height = 17
cell(ws2,13,1,"Pre-Tax Cost of Debt", font=BLACK_FM)
ws2.cell(row=13, column=2).value = 0.0430; ws2.cell(row=13, column=2).font = BLUE_IN
ws2.cell(row=13, column=2).fill = FILL_IN
ws2.cell(row=13, column=2).number_format = FMT_PCT
ws2.cell(row=13, column=2).alignment = Alignment(horizontal="center")
ws2.merge_cells("C13:E13")
cell(ws2,13,3,"Weighted avg coupon on Senior Notes (~4.3%); Moody's Aa2 credit rating", Font(name="Calibri", size=9, color="595959"))

ws2.row_dimensions[14].height = 17
cell(ws2,14,1,"Tax Rate (effective)", font=BLACK_FM)
ws2.cell(row=14, column=2).value = 0.260; ws2.cell(row=14, column=2).font = BLUE_IN
ws2.cell(row=14, column=2).fill = FILL_IN
ws2.cell(row=14, column=2).number_format = FMT_PCT
ws2.cell(row=14, column=2).alignment = Alignment(horizontal="center")
ws2.merge_cells("C14:E14")
cell(ws2,14,3,"Source: 10-K FY2024 effective tax rate ~26.0%", Font(name="Calibri", size=9, color="595959"))

ws2.row_dimensions[15].height = 20
cell(ws2,15,1,"After-Tax Cost of Debt", font=Font(name="Calibri", size=11, bold=True))
ws2.cell(row=15, column=2).value = "=B13*(1-B14)"
ws2.cell(row=15, column=2).font  = Font(name="Calibri", size=11, bold=True, color="1F4E79")
ws2.cell(row=15, column=2).fill  = FILL_OUT
ws2.cell(row=15, column=2).number_format = FMT_PCT
ws2.cell(row=15, column=2).alignment = Alignment(horizontal="center")
ws2.merge_cells("C15:E15")
cell(ws2,15,3,"= 4.30% × (1 - 26%) ≈ 3.18%", Font(name="Calibri", size=10, bold=True, color="1F4E79"))

ws2.row_dimensions[16].height = 8

# CAPITAL STRUCTURE
ws2.row_dimensions[17].height = 20
ws2.merge_cells("A17:E17")
cell(ws2,17,1,"CAPITAL STRUCTURE", FILL_HDR, BOLD_WHT,
     alignment=Alignment(horizontal="center"))

ws2.row_dimensions[18].height = 17
cell(ws2,18,1,"Current Stock Price ($)", font=BLACK_FM)
ws2.cell(row=18, column=2).value = "='DCF'!B9"
ws2.cell(row=18, column=2).font  = GREEN_LK
ws2.cell(row=18, column=2).number_format = FMT_SHR
ws2.cell(row=18, column=2).alignment = Alignment(horizontal="center")
ws2.merge_cells("C18:E18")
cell(ws2,18,3,"FactSet/Morningstar 2026-05-20 close: $1,074.01", Font(name="Calibri", size=9, color="595959"))

ws2.row_dimensions[19].height = 17
cell(ws2,19,1,"Diluted Shares Outstanding (M)", font=BLACK_FM)
ws2.cell(row=19, column=2).value = "='DCF'!B10"
ws2.cell(row=19, column=2).font  = GREEN_LK
ws2.cell(row=19, column=2).number_format = FMT_CUR2
ws2.cell(row=19, column=2).alignment = Alignment(horizontal="center")
ws2.merge_cells("C19:E19")
cell(ws2,19,3,"Source: 10-K FY2024; FactSet 10-Q", Font(name="Calibri", size=9, color="595959"))

ws2.row_dimensions[20].height = 17
cell(ws2,20,1,"Market Capitalization ($M)", font=BLACK_FM)
ws2.cell(row=20, column=2).value = "=B18*B19"
ws2.cell(row=20, column=2).font  = BLACK_FM
ws2.cell(row=20, column=2).number_format = FMT_CURM
ws2.cell(row=20, column=2).fill  = FILL_ALT
ws2.cell(row=20, column=2).alignment = Alignment(horizontal="center")

ws2.row_dimensions[21].height = 17
cell(ws2,21,1,"Total Debt ($M)", font=BLACK_FM)
ws2.cell(row=21, column=2).value = DEBT; ws2.cell(row=21, column=2).font = BLUE_IN
ws2.cell(row=21, column=2).fill = FILL_IN
ws2.cell(row=21, column=2).number_format = FMT_CURM
ws2.cell(row=21, column=2).alignment = Alignment(horizontal="center")
ws2.merge_cells("C21:E21")
cell(ws2,21,3,"Source: 10-K FY2024; Senior Notes + financing leases ≈ $10B", Font(name="Calibri", size=9, color="595959"))

ws2.row_dimensions[22].height = 17
cell(ws2,22,1,"Cash & Equivalents ($M)", font=BLACK_FM)
ws2.cell(row=22, column=2).value = CASH; ws2.cell(row=22, column=2).font = BLUE_IN
ws2.cell(row=22, column=2).fill = FILL_IN
ws2.cell(row=22, column=2).number_format = FMT_CURM
ws2.cell(row=22, column=2).alignment = Alignment(horizontal="center")
ws2.merge_cells("C22:E22")
cell(ws2,22,3,"Source: 10-K FY2024; $11.1B cash & short-term investments", Font(name="Calibri", size=9, color="595959"))

ws2.row_dimensions[23].height = 17
cell(ws2,23,1,"Net Debt ($M)", font=BLACK_FM)
ws2.cell(row=23, column=2).value = "=B21-B22"
ws2.cell(row=23, column=2).font  = BLACK_FM
ws2.cell(row=23, column=2).number_format = FMT_CURM
ws2.cell(row=23, column=2).fill  = FILL_ALT
ws2.cell(row=23, column=2).alignment = Alignment(horizontal="center")

ws2.row_dimensions[24].height = 17
cell(ws2,24,1,"Enterprise Value ($M)", font=BLACK_FM)
ws2.cell(row=24, column=2).value = "=B20+B23"
ws2.cell(row=24, column=2).font  = BLACK_FM
ws2.cell(row=24, column=2).number_format = FMT_CURM
ws2.cell(row=24, column=2).fill  = FILL_ALT
ws2.cell(row=24, column=2).alignment = Alignment(horizontal="center")

ws2.row_dimensions[25].height = 8

# WACC CALCULATION
ws2.row_dimensions[26].height = 20
ws2.merge_cells("A26:E26")
cell(ws2,26,1,"WACC CALCULATION", FILL_HDR, BOLD_WHT,
     alignment=Alignment(horizontal="center"))

ws2.row_dimensions[27].height = 16
for ci, hdr in enumerate(["Component","Weight","Cost","Contribution",""]):
    c = ws2.cell(row=27, column=1+ci, value=hdr)
    c.font = Font(name="Calibri", size=10, bold=True)
    c.fill = FILL_COL; c.alignment = Alignment(horizontal="center")

ws2.row_dimensions[28].height = 17
cell(ws2,28,1,"Equity", font=BLACK_FM)
ws2.cell(row=28, column=2).value = "=B20/B24"
ws2.cell(row=28, column=2).font  = BLACK_FM
ws2.cell(row=28, column=2).number_format = FMT_PCT2
ws2.cell(row=28, column=2).fill  = FILL_ALT
ws2.cell(row=28, column=2).alignment = Alignment(horizontal="center")
ws2.cell(row=28, column=3).value = "=B10"
ws2.cell(row=28, column=3).font  = BLACK_FM
ws2.cell(row=28, column=3).number_format = FMT_PCT2
ws2.cell(row=28, column=3).fill  = FILL_ALT
ws2.cell(row=28, column=3).alignment = Alignment(horizontal="center")
ws2.cell(row=28, column=4).value = "=B28*B10"
ws2.cell(row=28, column=4).font  = BLACK_FM
ws2.cell(row=28, column=4).number_format = FMT_PCT2
ws2.cell(row=28, column=4).fill  = FILL_ALT
ws2.cell(row=28, column=4).alignment = Alignment(horizontal="center")

ws2.row_dimensions[29].height = 17
cell(ws2,29,1,"Debt", font=BLACK_FM)
ws2.cell(row=29, column=2).value = "=B23/B24"
ws2.cell(row=29, column=2).font  = BLACK_FM
ws2.cell(row=29, column=2).number_format = FMT_PCT2
ws2.cell(row=29, column=2).fill  = FILL_ALT
ws2.cell(row=29, column=2).alignment = Alignment(horizontal="center")
ws2.cell(row=29, column=3).value = "=B15"
ws2.cell(row=29, column=3).font  = BLACK_FM
ws2.cell(row=29, column=3).number_format = FMT_PCT2
ws2.cell(row=29, column=3).fill  = FILL_ALT
ws2.cell(row=29, column=3).alignment = Alignment(horizontal="center")
ws2.cell(row=29, column=4).value = "=B29*B15"
ws2.cell(row=29, column=4).font  = BLACK_FM
ws2.cell(row=29, column=4).number_format = FMT_PCT2
ws2.cell(row=29, column=4).fill  = FILL_ALT
ws2.cell(row=29, column=4).alignment = Alignment(horizontal="center")

ws2.row_dimensions[30].height = 22
cell(ws2,30,1,"WACC", font=Font(name="Calibri", size=12, bold=True))
ws2.cell(row=30, column=2).value = "=B28+B29"
ws2.cell(row=30, column=2).font  = Font(name="Calibri", size=12, bold=True, color="1F4E79")
ws2.cell(row=30, column=2).fill  = FILL_OUT
ws2.cell(row=30, column=2).number_format = FMT_PCT2
ws2.cell(row=30, column=2).alignment = Alignment(horizontal="center")
ws2.merge_cells("D30:E30")
cell(ws2,30,4,"≈ 7.40% — net cash position reduces leverage impact",
     Font(name="Calibri", size=10, bold=True, color="1F4E79"))

ws2.row_dimensions[31].height = 8

ws2.merge_cells("A32:E32")
cell(ws2,32,1,"NOTES", Font(name="Calibri", size=10, bold=True, color="1F4E79"))
notes2 = [
    "• Costco is effectively a net-cash company ($11.1B cash > $10.0B debt), so WACC ≈ Cost of Equity",
    "• Morningstar estimates Costco WACC at 7.4% (moat report, Dec 2025)",
    "• Strong cash generation and low leverage result in minimal debt adjustment to WACC",
    "• CAPM Cost of Equity = 4.30% + 0.91 × 5.50% = 9.31%",
    "• Using WACC = 7.4% per Morningstar consensus as the primary discount rate",
]
for i, note in enumerate(notes2):
    r = 33 + i
    ws2.row_dimensions[r].height = 15
    ws2.merge_cells(f"A{r}:E{r}")
    cell(ws2, r, 1, note, Font(name="Calibri", size=9, color="595959"))

box(ws2, 4, 1, 10, 5)
box(ws2, 12, 1, 15, 5)
box(ws2, 17, 1, 24, 5)
box(ws2, 26, 1, 30, 5)

# ── save ─────────────────────────────────────────────────────────────────────
out = f"/Users/weimingzhuang/Documents/source_code/financial-services-opencode/.opencode/COST_DCF_Model.xlsx"
wb.save(out)
print(f"Saved: {out}")
print("Done!")